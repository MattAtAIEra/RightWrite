"""
Parse all vocabulary Excel files from resource/ and generate backend/vocab_all.json.
Uses pypinyin to generate similar_wrong (homophone) lists for each character.
Preserves manually curated similar_wrong from existing vocab_data.py where available.
"""
import json
import os
import re
import sys

from openpyxl import load_workbook
from pypinyin import pinyin, Style

RESOURCE_DIR = os.path.join(os.path.dirname(__file__), "..", "resource")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "backend", "vocab_all.json")

PUBLISHER_MAP = {
    "康軒版": "kangxuan",
    "南一版": "nanyi",
    "翰林版": "hanlin",
}
PUBLISHER_LABELS = {"kangxuan": "康軒版", "nanyi": "南一版", "hanlin": "翰林版"}
GRADE_CN = {1: "一", 2: "二", 3: "三", 4: "四", 5: "五", 6: "六"}

# Regex to extract lesson number from directory/filename
LESSON_NUM_RE = re.compile(r"第([一二三四五六七八九十百零]+)課")
CN_NUM = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9, "十": 10}


def cn_to_int(s: str) -> int:
    """Convert Chinese number string to int (handles 1-20)."""
    if len(s) == 1:
        return CN_NUM.get(s, 0)
    if s.startswith("十"):
        return 10 + CN_NUM.get(s[1:], 0)
    if s.endswith("十"):
        return CN_NUM.get(s[0], 0) * 10
    return 0


def is_cjk(ch: str) -> bool:
    return len(ch) == 1 and "\u4e00" <= ch <= "\u9fff"


def extract_examples(text: str) -> list[str]:
    """Extract example sentences from cell text. Examples follow [���] markers."""
    if not text:
        return []
    examples = []
    for m in re.finditer(r"[【\[（(]例[】\]）)](.+?)(?=[。！？]|$)", text, re.DOTALL):
        sent = m.group(1).strip().rstrip("。！？") + "。"
        if len(sent) > 3:
            examples.append(sent)
    # Also try lines that start after [例]
    parts = re.split(r"[【\[（(]例[】\]）)]", text)
    if len(parts) > 1:
        for part in parts[1:]:
            for sent in re.split(r"[。！？]", part):
                sent = sent.strip()
                if len(sent) > 2:
                    examples.append(sent + "。")
    return list(dict.fromkeys(examples))[:3]  # deduplicate, max 3


def parse_excel(filepath: str) -> dict:
    """Parse a single vocabulary Excel file. Returns lesson data dict."""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    # Extract title: prefer from cell A1, fallback to filename
    title_cell = ws.cell(1, 1).value or ""
    title = title_cell.strip()
    if not title:
        # Get title from filename: "第一課：一束鮮花.xlsx" → "一束鮮花"
        fname = os.path.splitext(os.path.basename(filepath))[0]
        title = fname

    # Extract clean title (after ：or :)
    if "：" in title:
        title = title.split("：", 1)[1].strip()
    elif ":" in title:
        title = title.split(":", 1)[1].strip()

    characters = []
    compounds = []

    # Skip row 1 (title) and row 2 (headers)
    for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
        word = (row[0] or "").strip()
        definition = (row[1] or "").strip() if row[1] else ""

        if not word:
            continue

        examples = extract_examples(definition)

        if len(word) == 1 and is_cjk(word):
            characters.append({
                "char": word,
                "similar_wrong": [],  # will be filled later
                "examples": examples,
            })
        elif len(word) > 1:
            compounds.append({
                "word": word,
                "examples": examples,
            })

    wb.close()
    return {"title": title, "characters": characters, "compounds": compounds}


def load_existing_similar_wrong() -> dict[str, list[str]]:
    """Load manually curated similar_wrong from saved JSON."""
    curated_path = os.path.join(os.path.dirname(__file__), "curated_similar_wrong.json")
    if os.path.exists(curated_path):
        with open(curated_path, encoding="utf-8") as f:
            curated = json.load(f)
        return curated
    print("Warning: curated_similar_wrong.json not found, generating all from pypinyin")
    return {}


def build_pinyin_map(all_chars: set[str]) -> dict[str, list[str]]:
    """Build pinyin -> characters mapping for homophone lookup."""
    py_map: dict[str, list[str]] = {}
    py_no_tone_map: dict[str, list[str]] = {}

    for ch in sorted(all_chars):
        pys = pinyin(ch, style=Style.TONE3, heteronym=False)
        if pys and pys[0]:
            py_str = pys[0][0]  # e.g., "jie2"
            py_map.setdefault(py_str, []).append(ch)
            # Also map without tone number
            base = re.sub(r"\d$", "", py_str)
            py_no_tone_map.setdefault(base, []).append(ch)

    return py_map, py_no_tone_map


def generate_similar_wrong(char: str, py_map: dict, py_no_tone_map: dict, all_chars: set) -> list[str]:
    """Generate similar_wrong list for a character using pinyin homophones."""
    pys = pinyin(char, style=Style.TONE3, heteronym=False)
    if not pys or not pys[0]:
        return []

    py_str = pys[0][0]
    base = re.sub(r"\d$", "", py_str)

    # Priority 1: exact pinyin match (same tone)
    candidates = [c for c in py_map.get(py_str, []) if c != char]

    # Priority 2: same base pinyin (different tone)
    if len(candidates) < 2:
        for c in py_no_tone_map.get(base, []):
            if c != char and c not in candidates:
                candidates.append(c)

    # Limit to 4
    return candidates[:4]


def scan_all_excels() -> dict:
    """Scan resource/ directory and build the complete dataset."""
    curated = load_existing_similar_wrong()
    print(f"Loaded {len(curated)} curated similar_wrong entries")

    # First pass: collect all characters and parse all files
    all_grades = {}
    all_chars = set()

    for dir_name in sorted(os.listdir(RESOURCE_DIR)):
        dir_path = os.path.join(RESOURCE_DIR, dir_name)
        if not os.path.isdir(dir_path) or dir_name.startswith("."):
            continue

        # Parse directory name: "四下-康軒版"
        match = re.match(r"([一二三四五六])下-(.+版)", dir_name)
        if not match:
            print(f"Skipping unrecognized directory: {dir_name}")
            continue

        grade_cn, publisher = match.groups()
        grade_num = CN_NUM.get(grade_cn, 0)
        pub_code = PUBLISHER_MAP.get(publisher)
        if not grade_num or not pub_code:
            print(f"Skipping unknown grade/publisher: {dir_name}")
            continue

        grade_id = f"{grade_num}_{pub_code}"
        lessons = {}

        for fname in sorted(os.listdir(dir_path)):
            if not fname.endswith(".xlsx") or fname.startswith("~"):
                continue
            # Skip duplicate files like "第一課：xxx (1).xlsx"
            if re.search(r"\(\d+\)\.xlsx$", fname):
                continue

            lesson_match = LESSON_NUM_RE.search(fname)
            if not lesson_match:
                continue
            lesson_num = cn_to_int(lesson_match.group(1))
            if lesson_num == 0:
                continue

            filepath = os.path.join(dir_path, fname)
            try:
                lesson_data = parse_excel(filepath)
                lessons[lesson_num] = lesson_data
                for ch in lesson_data["characters"]:
                    all_chars.add(ch["char"])
            except Exception as e:
                print(f"Error parsing {filepath}: {e}")

        if lessons:
            total = len(lessons)
            mid = total // 2
            all_grades[grade_id] = {
                "label": dir_name,
                "semester": "114學年度第2學期",
                "grade": f"{grade_cn}年級",
                "publisher": publisher,
                "total_lessons": total,
                "midterm_range": [1, mid],
                "final_range": [mid + 1, total],
                "lessons": lessons,
            }

    print(f"Collected {len(all_chars)} unique characters from {len(all_grades)} grade/publisher combos")

    # Build pinyin maps
    py_map, py_no_tone_map = build_pinyin_map(all_chars)

    # Second pass: fill similar_wrong
    for grade_id, grade_data in all_grades.items():
        for lesson_num, lesson in grade_data["lessons"].items():
            for ch_data in lesson["characters"]:
                char = ch_data["char"]
                if char in curated:
                    ch_data["similar_wrong"] = curated[char]
                else:
                    ch_data["similar_wrong"] = generate_similar_wrong(
                        char, py_map, py_no_tone_map, all_chars
                    )

    return all_grades


def main():
    data = scan_all_excels()

    # Stats
    total_chars = 0
    chars_with_sw = 0
    for gd in data.values():
        for lesson in gd["lessons"].values():
            for ch in lesson["characters"]:
                total_chars += 1
                if ch["similar_wrong"]:
                    chars_with_sw += 1

    print(f"\nTotal characters: {total_chars}")
    print(f"Characters with similar_wrong: {chars_with_sw} ({100*chars_with_sw//total_chars}%)")
    print(f"Output: {OUTPUT_PATH}")

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print("Done!")


if __name__ == "__main__":
    main()
